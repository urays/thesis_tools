import os
import openpyxl
import pytesseract
from PIL import Image

curpath = os.path.abspath(os.path.dirname(__file__))
for root,dirs,files in os.walk(curpath):
    for dir in dirs:
        abspath = os.path.join(root, dir)
        filenames = []
        value = []
        for x in os.listdir(abspath):
            if x[-3:] == 'png':
                filenames.append(x[:-4])
                image = Image.open(abspath+"\\"+x)
                num = pytesseract.image_to_string(image)
                value.append(num)

        xlsx = openpyxl.Workbook()
        table = xlsx.active
        for i in range(0, len(filenames)):
            table.cell(row=i + 1, column=1, value=i+1)
            table.cell(row=i + 1, column=2, value=filenames[i])
            if value[i]!='':
                table.cell(row=i + 1, column=3, value=int(value[i]))
        xlsx.save(dir+'.xlsx')
    break




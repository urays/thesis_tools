# -*- encoding: utf-8 -*-
# extract keywords from papers

import os
import openpyxl
from extpdf import extract_keywords


exkws = extract_keywords(maxpages=2,
                         s_marks=["eywords—", "eywords-", "erms—", "erms-"],
                         e_marks=[['I.', 0]],
                         sepchar=[';', ','],
                         region=150,
                         maxchars=50,
                         maxcnt=10)

PATH = ".\\2019"
xlsx_name = "keywords.xlsx"

pns = []
for file in os.listdir(PATH):
    fp = os.path.join(PATH, file)
    a, b = os.path.split(fp)
    sn, ext = os.path.splitext(b)
    if ext == ".pdf":
        ws = exkws.search(paper=fp)
        # print(ws)
        pns.append([sn, ws])

for x in pns:
    print("[paper] =>", x[0])
    print("[keywords] =>", x[1], end="\n\n")

if os.path.exists(xlsx_name):
    os.remove(xlsx_name)

xlsx = openpyxl.Workbook()
table = xlsx.active
for i in range(0, len(pns)):
    if len(pns[i]) == 0:
        continue
    table.cell(row=i + 1, column=1, value=str(pns[i][0]))
    for j in range(len(pns[i][1])):
        table.cell(row=i + 1, column=(j + 2), value=str(pns[i][1][j]))
    print("=>", pns[i][0])
xlsx.save(xlsx_name)

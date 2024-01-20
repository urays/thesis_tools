# -*- encoding: utf-8 -*-

import os

import openpyxl
import requests
from bs4 import BeautifulSoup


class Top500:
    def __init__(self):
        self.url = "https://www.top500.org/lists/top500/"
        self.heads = [
            "Rank", "System", "Site", "Cores", "Rmax (TFlop/s)",
            "Rpeek (TFlop/s)", "Power(kW)"
        ]

    def getTop(self):
        response = requests.get(self.url)
        soup = BeautifulSoup(response.text, 'html.parser')
        squares = soup.find('ul', {"id": "squarelist"})
        links = squares.find_all('a')
        listaComparativas = []
        for ref in links:
            save = str(ref['href']).replace('lists', 'list')
            listaComparativas.append(save)

        return listaComparativas

    def getDates(self, dates=["2020/06"], n=5):
        datas = {}
        for p in dates:
            datas[p] = []
            datas[p].append(self.heads)

            print("=>", p)
            response = requests.get(self.url + p)
            soup = BeautifulSoup(response.text, 'html.parser')
            table = soup.find('table', {'class': "table table-condensed table-striped"})
            body = table.find_all('tr')

            usol = 0
            cout = 0
            for b in body:
                if cout >= n:
                    break
                if usol == 0:
                    usol = 1
                    continue
                b = b.find_all("td")
                system = b[1].find('a').find('b')
                if system is None:
                    system = b[1].find('a').text.split(',')[0]
                else:
                    system = system.text
                site = str(b[1]).replace("<br/>", "<br>").split("<br>")[2][:-20].rstrip()
                example = [b[0].text,  # rank
                           system,  # system
                           site,  # site
                           int(b[2].text.replace(',', '')),   # cores
                           float(b[3].text.replace(',', '')),  # rmax
                           float(b[4].text.replace(',', '')),  # rpeek
                           int(b[5].text.replace(',', '')) if b[5].text != "" else None,  # power
                           ]
                cout += 1
                # print(example)
                datas[p].append(example)
        return datas


if __name__ == '__main__':
    dd = Top500()
    date = [
        "2020-06", "2019-11", "2019-06", "2018-11", "2018-06", "2017-11",
        "2017-06", "2016-11", "2016-06", "2015-11", "2015-06"
    ]
    # date = ["2020-06", "2015-06"]
    datas = dd.getDates(dates=[x.replace('-', '/') for x in date], n=5)
    # print(datas)

    xlsx_name = "top500.xlsx"
    if os.path.exists(xlsx_name):
        os.remove(xlsx_name)

    # write into xlsx
    xlsx = openpyxl.Workbook()
    [xlsx.create_sheet(x) for x in date]
    for k in date:
        table = xlsx.get_sheet_by_name(k)
        dt = datas[k.replace('-', '/')]
        for i in range(0, len(dt)):
            for j in range(0, len(dt[i])):
                table.cell(row=i + 1, column=j + 1, value=str(dt[i][j]))
        print("=>", xlsx_name, "<", k, "ok")
    xlsx.save(xlsx_name)
